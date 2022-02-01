import math
from pathlib import Path
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import openpyxl
from openpyxl.workbook.workbook import Workbook
import xlsxwriter as xw
import os
from tkinter import filedialog

class Calculate:
    def __init__(self, obj):
        self.obj = obj

    '''
        method: calculate all the sheet value

        target_file: already created file with values
                     we wil retreive all the value to
                     calculate pure value
        new file: which we will create using target file
                  calculate value
    '''
        
    def calculate_purity(self):
        # getting current date
        self.date = self.obj.entry7.get()

        # getting the target file path from tkinter entry
        self.path_excel_sheet = self.obj.entry0.get()

        # saved file location
        self.path_saved_location = self.obj.entry1.get()
        # if file path is specified than
        # file will save on specified location
        # otherwise it will save on default loaction (Desktop/current_date/file.xlsx)
        if self.path_saved_location == "":
            self.file_saved_location(r"C:\Users\lenovo\Desktop\PurityCalculatorFiles")
        else:
            self.file_saved_location(f"{self.path_saved_location}\PurityCalculatorFiles")

        # getting target file sheet name
        self.sheet_name = self.obj.entry2.get()

        # getting target file raw parcent positon
        self.raw_percent_position = self.obj.entry3.get()

        # getting target file weight position
        self.raw_weight_position = self.obj.entry4.get()

        # getting target file raw position
        # where number is start
        # after header
        self.raw_position = int(self.obj.entry5.get())

        # substraction value
        self.value = self.obj.entry6.get()

        # method: creating new excel file
        self.create_xlsx_file()

        # method: retrieve data from target file
        self.open_target_sheet()

        # method: adding default data to new file
        #         addind column header
        self.adding_default_data_to_xlsx()

        # method: adding calculated value to the
        #         new xlsx file
        self.adding_value_to_xlsx()

    '''
        location of new xlsx file
        if input_path_is_empty:
            create_new_path
        else:
            default_path
    '''
    def file_saved_location(self, location):
        self.newPath = location + f"\{self.date}"

        # if folder_exists:
        #   saved_file_on_available_path
        # else:
        #   create_new_path 
        if not os.path.exists(self.newPath):
            os.makedirs(self.newPath)

    '''
        create new xlsx file using
        xlswriter
    '''
    def create_xlsx_file(self):
        # retreiving target path
        self.retrieving_file_path = Path(self.path_excel_sheet)
        # create a workbook 
        # save the file
        # named [location/(substract_value)(target_file_name)]
        self.create_xlsx = xw.Workbook(f"{self.newPath}\(-{self.value}){os.path.basename(self.retrieving_file_path)}")
        # adding new sheet to the new xlsx file
        self.add_sheet_to_xlsx = self.create_xlsx.add_worksheet()
        print("calc: create_xlsx_file")

    '''
        opening the target sheet using openpyxl
        getting sheet name
    '''
    def open_target_sheet(self):
        # open target sheet
        self.open_sheet_from_path = load_workbook(f'{self.path_excel_sheet}')
        # getting target sheet
        self.get_sheet = self.open_sheet_from_path[self.sheet_name]
        print("calc: open_target_sheet")

    '''
        adding default value to new xlsx file
    '''
    def adding_default_data_to_xlsx(self):
        self.add_sheet_to_xlsx.write("B1","Purity")
        self.add_sheet_to_xlsx.write("C1",f"Purity(-{self.value})")
        self.add_sheet_to_xlsx.write("D1","Weight")
        self.add_sheet_to_xlsx.write("E1","Pure Weight")
        print("calc: adding_default_data_to_xlsx")

    '''
        getting each value from target file
        adding the value to the new xlsx file
    '''
    def adding_value_to_xlsx(self):
        # total raw on target sheet
        self.sheet_raw_count = self.get_sheet.max_row + 1

        # var: total impure gold
        total_impure = 0
        # var: total pure gold
        total_pure = 0

        # iteration for getting all the values and add values
        # it will starts from raw_postion(where numerical value starts
        #                                 on target file)
        # ends sheet_raw_count(maximum raw on target sheet)
        for i in range(self.raw_position, self.sheet_raw_count):
            # parcent raw on target sheet
            # getting positon (column A/B/C) from tkinter entry(entry3)
            get_percent_raw = self.get_sheet[f'{self.raw_percent_position}{i}']
            # weight raw on (column A/B/C) target sheet
            # getting position from tkinter entry(entry4)
            get_weight_raw = self.get_sheet[f'{self.raw_weight_position}{i}']
            # get parcent value
            percent_value = get_percent_raw.value
            # get weight value
            weight_value = get_weight_raw.value

            # -------------- Formula ----------------
            # pure_value = ((percent - .50) * weight)%
            # step-1: less_value = percent - input_value(.20/.50)
            # step-2: pure_value = (less_value * weight)/100
            
            # step-1
            less_value = float(percent_value) - float(self.value)
            # step-2
            pure_value = (less_value * weight_value)/100

            # write/add the value to the new file
            self.add_sheet_to_xlsx.write(f"A{i+1}", i)
            self.add_sheet_to_xlsx.write(f"B{i+1}", percent_value)
            self.add_sheet_to_xlsx.write(f"C{i+1}", less_value)
            self.add_sheet_to_xlsx.write(f"D{i+1}", weight_value)
            self.add_sheet_to_xlsx.write(f"E{i+1}", self.truncate(pure_value, 2))
            
            # getting all impure gold
            # impure gold: all target sheet weight value
            # getting sum by: adding each value on every itaration
            total_impure += weight_value
            # getting all pure gold
            # pure gold: pure weight value after calculation
            # getting sum by: adding each value on every itaration
            total_pure += self.truncate(pure_value, 2)


        # write/add impure_gold and pure gold to new sheet
        self.add_sheet_to_xlsx.write(f"D{self.sheet_raw_count + 2}", total_impure)
        self.add_sheet_to_xlsx.write(f"E{self.sheet_raw_count + 2}", total_pure)
        # closing the new xlsx file
        self.create_xlsx.close()

    # if value is
    # > weight: 78.879
    # after floor
    # > return: 78.88
    # but we need 
    # > return: 78.87
    # created the method to solve this problem
    def truncate(self, f, n):
        return math.floor(f*10**n) / 10**n

    # def print_excel_sheet_hard_copy(self):

    #     self.path = filedialog.askopenfilename()
    #     # loading the Excel File and the sheet
    #     self.load_file = openpyxl.load_workbook(self.path)
    #     self.load_file_sheet = self.load_file["Sheet1"]

    #     # calling the image_loader
    #     self.img_loader = SheetImageLoader(self.load_file_sheet)

    #     # get the image
    #     self.img = self.img_loader.get("A1")

    #     # showing the image
    #     self.img.show()

    #     # saving the image
    #     self.img.save('sheet.jpg')


            






        

